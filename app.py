# LGePR Data Cleaner v12.7 (Multiple Export Files)
# NEW: 3 pliki do pobrania po merge:
#   - LGePR_FINAL (bez zmian)
#   - LGePR_raw_data (Uploaded, Published, Media, Media Type, Media Grade, Headline, Country, Division, Product, PR Value, ESG, M/Z)
#   - Reach (Division, Product, Medium, Source, Title.pl, Title.eng, Reach, Date, ESG, MZ)
# NEW: Merge przepisuje dodatkowe kolumny z raportu (Uploaded, Published, Media Type, Media Grade, Country)

import streamlit as st
import pandas as pd
import re
import io
import time
import json
import os
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl

# Biblioteki AgGrid
try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode, DataReturnMode
    AGGRID_AVAILABLE = True
except ImportError:
    AGGRID_AVAILABLE = False

# Próba importu newspaper
try:
    from newspaper import Article
    NEWSPAPER_AVAILABLE = True
except ImportError:
    NEWSPAPER_AVAILABLE = False

# ─────────────────────────────────────────────
# 1. KONFIGURACJA
# ─────────────────────────────────────────────
st.set_page_config(page_title="LGePR Cleaner", page_icon="🧹", layout="wide")

hide_ui_css = """
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
.stDeployButton {display:none;}
div[data-testid="stDecoration"] {display:none;}
</style>
"""
st.markdown(hide_ui_css, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 2. DEFINICJE I REGUŁY
# ─────────────────────────────────────────────
TITLE_MAX_LEN = 120
QUOTE_MAX_LEN = 120
ID_TITLE_CHARS = 30

SANITIZATION_PATTERN = re.compile(r'[.:!?"\'()\[\]/\\$€£zł\-–—,;]')
YEAR_PATTERN = re.compile(r'\b2026\b')

FINAL_OUTPUT_ORDER = [
    'dim1', 'dim2', 'Print', 'source', 'title', 'AVE [PLN]', 'reach', 'date', 
    'unique users', 'ENG Title', 'Division', 'Product', 'Clipping', 'LGePR',
    'ESG', 'M/Z', 'Links', 'PR Value', 'Quote', 'LG', 'Exclusive', 'Photo'
]

VALIDATION_RULES = {
    "Division": ["Corporate", "HS", "MS", "VS", "ES"],
    "Photo": ["None", "LGE logo", "product", "personnel"],
    "Exclusive": ["Exclusive", "66"],
    "LG": ["N/A", "LG Electronics"],
    "ESG": ["Yes", "No"],
    "M/Z": ["Yes", "No"]
}

PRODUCT_RULES = {
    "Corporate": ["Corporate/Brand", "Top Management", "Finance", "MC", "Others"],
    "HS": ["Refrigerator", "Washer/Dryer", "Cooking Appliance", "Vacuum Cleaner", "Styler", "Air Solution", "Others"],
    "MS": ["LCD TV", "Audio", "OLED TV", "Signage", "PC", "Projector", "Monitor", "Others"],
    "VS": ["VS"],
    "ES": ["SAC", "RAC", "AirCare", "Chiller", "Others"]
}

# ─────────────────────────────────────────────
# 3. OBSŁUGA SEKRETÓW
# ─────────────────────────────────────────────
def get_secret(key, default=None):
    try: return st.secrets.get(key, default)
    except: return default

def check_password():
    if "password_correct" not in st.session_state:
        st.session_state.password_correct = False
    if st.session_state.password_correct:
        return True
    
    st.markdown("### 🔒 Dostęp autoryzowany")
    pwd = st.text_input("Hasło:", type="password")
    if st.button("Zaloguj"):
        secret_pwd = get_secret("APP_PASSWORD", "admin123")
        if pwd == secret_pwd:
            st.session_state.password_correct = True
            st.rerun()
        else:
            st.error("Błędne hasło")
    return False

if not check_password():
    st.stop()

def get_cloud_config():
    api_key = get_secret("OPENAI_API_KEY", "")
    raw_media_list = get_secret("MEDIA_LIST", [])
    if isinstance(raw_media_list, str): 
        raw_media_list = [x.strip() for x in raw_media_list.split(',')]
    
    normalized_set = set()
    for m in raw_media_list:
        clean_m = normalize_domain(m).lower()
        if clean_m: normalized_set.add(clean_m)
    return api_key, normalized_set

# ─────────────────────────────────────────────
# 4. POMOCNIKI
# ─────────────────────────────────────────────
def normalize_domain(val):
    if pd.isna(val): return ""
    val_str = str(val).strip()
    if '.' not in val_str: return val_str
    
    u = val_str.lower()
    u = re.sub(r'^https?://', '', u)
    u = re.sub(r'^www\.', '', u)
    if u.endswith('/'): u = u[:-1]
    
    if u.endswith('.onet.pl') or u == 'onet.pl': return 'onet.pl'
    if u.endswith('.wp.pl') or u == 'wp.pl': return 'wp.pl'
    if u.endswith('.gazeta.pl') or u == 'gazeta.pl': return 'gazeta.pl'
    if u.endswith('.interia.pl') or u == 'interia.pl': return 'interia.pl'
    if u.endswith('.infor.pl') or u == 'infor.pl': return 'infor.pl'
    if u.endswith('.rp.pl') or u == 'rp.pl': return 'rp.pl'
    
    mapping = {
        'komputerswiat.pl': 'onet.pl', 'auto-swiat.pl': 'onet.pl', 'businessinsider.com.pl': 'onet.pl', 'plejada.pl': 'onet.pl', 'medonet.pl': 'onet.pl', 'forbes.pl': 'onet.pl',
        'benchmark.pl': 'wp.pl', 'gadzetomania.pl': 'wp.pl', 'dobreprogramy.pl': 'wp.pl', 'pudelek.pl': 'wp.pl', 'money.pl': 'wp.pl', 'autokult.pl': 'wp.pl',
        'next.gazeta.pl': 'gazeta.pl', 'sport.pl': 'gazeta.pl', 'plotek.pl': 'gazeta.pl', 'moto.pl': 'gazeta.pl',
        'pomponik.pl': 'interia.pl', 'swiatseriali.interia.pl': 'interia.pl'
    }
    return mapping.get(u, u)

def has_value(val):
    if val is None: return False
    try:
        if pd.isna(val): return False
    except: pass
    s = str(val).strip()
    # Uwaga: "None" NIE jest tutaj bo dla kolumny Photo to valid value
    if s == "" or s.lower() in ["nan", "[no_content]", "[ai_fail]", "[json_err]", "[no_img]", "error getting image"] or "error" in s.lower(): 
        return False
    return True

def validate_val(val, allowed_list):
    if not has_value(val): return False
    return str(val).strip() in [str(x) for x in allowed_list]

def enforce_strict_rules(key, value, context_division=None):
    val_str = str(value).strip()
    if key == "Division":
        if val_str in VALIDATION_RULES["Division"]: return val_str
        return "[CHECK]"
    if key == "Product":
        allowed = []
        if context_division and context_division in PRODUCT_RULES:
            allowed = PRODUCT_RULES[context_division]
        else:
            for p_list in PRODUCT_RULES.values(): allowed.extend(p_list)
        if val_str in allowed: return val_str
        if "LED TV" in val_str: return "LCD TV"
        if "Vrand" in val_str: return "Corporate/Brand"
        return "Others" if "Others" in allowed else "[CHECK]"
    if key == "Photo":
        return val_str if val_str in VALIDATION_RULES["Photo"] else "None"
    if key == "Exclusive":
        return val_str if str(val_str) in [str(x) for x in VALIDATION_RULES["Exclusive"]] else "Exclusive"
    return val_str

def clean_text(t, l):
    if pd.isna(t): return ""
    x = str(t).strip()
    x = YEAR_PATTERN.sub("2026r", x)
    x = SANITIZATION_PATTERN.sub(" ", x)
    x = re.sub(r'\s+', ' ', x).strip()
    if len(x) > l:
        x = x[:l]
        last_space = x.rfind(' ')
        if last_space != -1: x = x[:last_space]
    return x.strip()

def scrape_article_data(url):
    if not str(url).startswith('http'): url = 'https://' + str(url)
    result = {"text": "", "image_url": None}
    
    if NEWSPAPER_AVAILABLE:
        try:
            a = Article(url)
            a.download()
            a.parse()
            result["text"] = a.text[:4000] if a.text else ""
            result["image_url"] = a.top_image
            if result["text"]: return result
        except: pass

    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.content, 'html.parser')
        paragraphs = soup.find_all('p')
        text_content = " ".join([p.get_text() for p in paragraphs])
        
        image_url = None
        meta_img = soup.find('meta', property='og:image')
        if meta_img: image_url = meta_img.get('content')
        
        result["text"] = text_content[:4000] if text_content else ""
        if not result["image_url"]: result["image_url"] = image_url
    except: pass
    
    return result

def extract_specific_columns(f, sheet, media_list_set) -> pd.DataFrame:
    wb = openpyxl.load_workbook(f, data_only=False)
    ws = wb[sheet]
    headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column+1) if ws.cell(1, c).value}
    
    data = []
    for r in range(2, ws.max_row+1):
        src_val = ws.cell(r, headers.get('source', 4)).value
        tit_val = ws.cell(r, headers.get('title', 5)).value
        rea_val = ws.cell(r, headers.get('reach', 7)).value
        dat_val = ws.cell(r, headers.get('date of service', 8)).value
        div_val = ws.cell(r, headers.get('Division', 10)).value
        prod_val = ws.cell(r, 11).value
        excl_val = ws.cell(r, 12).value
        phot_val = ws.cell(r, 13).value
        
        # Wyciągnij hiperłącze
        link = ""
        c = ws.cell(r, headers.get('source', 4))
        if c.hyperlink and c.hyperlink.target: 
            link = c.hyperlink.target
        elif isinstance(c.value, str) and c.value.startswith('http'): 
            link = c.value
        
        # Znormalizowana domena dla kolumny 'source'
        clean_src_display = normalize_domain(src_val)
        
        # Sprawdzenie czy media jest na liście
        check_val = clean_src_display.lower()
        stat = "OK" if media_list_set and check_val in media_list_set else "BRAK"
        if not media_list_set: stat = "N/A"
        
        # LG w tytule
        lg_calc = "LG Electronics" if "LG" in str(tit_val).upper() else "N/A"
        
        # Data - tylko dzień
        day = str(dat_val)
        try: 
            day = str(pd.to_datetime(dat_val).day)
        except: 
            pass

        # Print: "Yes" jeśli brak linku, puste jeśli jest link
        print_val = "Yes" if not link else ""
        
        # Pełny link bez http:// dla kolumny Links
        link_clean = re.sub(r'^https?://', '', str(link).strip()) if link else ""

        row = {
            'dim1': "",
            'dim2': "",
            'Print': print_val,
            'source': clean_src_display,
            'title': tit_val,  # Oryginalny tytuł PL
            'AVE [PLN]': "",
            'reach': rea_val,
            'date': day,
            'unique users': "",
            'ENG Title': "",  # Będzie uzupełnione przez AI
            'Division': div_val,
            'Product': prod_val,
            'Clipping': "",
            'LGePR': "",
            'ESG': "",
            'M/Z': "",
            'Links': link_clean,
            'PR Value': "",
            'Quote': "",  # Będzie uzupełnione przez AI
            'LG': lg_calc,
            'Exclusive': excl_val,
            'Photo': phot_val,
            # Wewnętrzne kolumny (nie będą w finalnym eksporcie)
            '_orig_date': dat_val,
            '_media_status': stat
        }
        data.append(row)
    wb.close()
    return pd.DataFrame(data)

def generate_id_match(row):
    src = str(row.get('source', '')).strip()
    # Używamy ENG Title jeśli jest, w przeciwnym razie title
    eng_title = row.get('ENG Title', '')
    orig_title = row.get('title', '')
    tit = str(eng_title if has_value(eng_title) else orig_title)[:ID_TITLE_CHARS].strip()
    try: 
        d = pd.to_datetime(row.get('_orig_date')).strftime("%Y%m%d")
    except: 
        d = str(row.get('_orig_date', ''))[:8].replace('-','')
    return f"{src}|{tit}|{d}"

def merge_datasets(clean_df, report_df):
    """
    Łączy plik Clean z raportem PR Value.
    
    Klucz: source/Media + ENG Title/Headline + date/Published (dzień)
    Fallback: gdy subdomeny różne ale główna domena ta sama (np. tech.wp.pl vs wp.pl)
    """
    
    # 1. Autodetekcja kolumn w Raporcie
    media_col = 'Media'
    if 'Media' not in report_df.columns:
        if 'Source' in report_df.columns: media_col = 'Source'
    
    title_col = 'Headline'
    if 'Headline' not in report_df.columns:
        if 'Title' in report_df.columns: title_col = 'Title'
        elif 'Tytuł' in report_df.columns: title_col = 'Tytuł'
    
    date_col = 'Published'
    if 'Published' not in report_df.columns:
        if 'Data' in report_df.columns: date_col = 'Data'
        elif 'Date' in report_df.columns: date_col = 'Date'
        
    pr_col = 'PR Value'
    if 'PR Value' not in report_df.columns:
        if 'AVE' in report_df.columns: pr_col = 'AVE'

    # 2. Konwertuj PR Value na liczby (usuwając spacje jako separator tysięcy)
    def parse_pr_value(val):
        if pd.isna(val):
            return 0
        if isinstance(val, (int, float)):
            return val
        # String ze spacjami jako separator tysięcy, np. "1 021" lub "126 229"
        val_str = str(val).replace(' ', '').replace('\xa0', '').strip()
        try:
            return float(val_str)
        except:
            return 0
    
    report_df[pr_col] = report_df[pr_col].apply(parse_pr_value)

    # 3. Funkcja wyciągająca główną domenę (np. tech.wp.pl -> wp.pl)
    def get_main_domain(domain):
        domain = normalize_domain(str(domain)).lower()
        if not domain:
            return ""
        
        # Lista znanych głównych domen z subdomenami
        known_mains = ['wp.pl', 'onet.pl', 'gazeta.pl', 'interia.pl', 'rp.pl', 'infor.pl']
        
        for main in known_mains:
            if domain == main or domain.endswith('.' + main):
                return main
        
        # Dla innych - zwróć całą domenę (bez zmian)
        return domain

    # 4. Funkcja generująca klucz łączenia
    def create_key(media_val, title_val, date_val):
        import numpy as np
        
        # Media - normalizuj domenę
        m = normalize_domain(str(media_val)).lower()
        
        # Tytuł - wyczyść i weź pierwsze 30 znaków
        t_clean = clean_text(str(title_val), 200) 
        t = t_clean.lower().strip()[:30]
        
        # Data - wyciągnij tylko dzień
        if isinstance(date_val, (int, float, np.integer, np.floating)) and not pd.isna(date_val):
            d = str(int(date_val))
        else:
            try:
                d = str(pd.to_datetime(date_val).day)
            except:
                d = str(date_val).strip()
        
        return f"{m}|{t}|{d}"
    
    # 5. Funkcja generująca klucz z główną domeną (fallback)
    def create_key_main_domain(media_val, title_val, date_val):
        import numpy as np
        
        # Media - główna domena
        m = get_main_domain(media_val)
        
        # Tytuł - wyczyść i weź pierwsze 30 znaków
        t_clean = clean_text(str(title_val), 200) 
        t = t_clean.lower().strip()[:30]
        
        # Data - wyciągnij tylko dzień
        if isinstance(date_val, (int, float, np.integer, np.floating)) and not pd.isna(date_val):
            d = str(int(date_val))
        else:
            try:
                d = str(pd.to_datetime(date_val).day)
            except:
                d = str(date_val).strip()
        
        return f"{m}|{t}|{d}"

    # 6. Generowanie kluczy dla CLEAN
    clean_df['__merge_key'] = clean_df.apply(
        lambda r: create_key(
            r['source'], 
            r['ENG Title'] if has_value(r.get('ENG Title')) else r.get('title', ''), 
            r['date']
        ), 
        axis=1
    )
    clean_df['__merge_key_main'] = clean_df.apply(
        lambda r: create_key_main_domain(
            r['source'], 
            r['ENG Title'] if has_value(r.get('ENG Title')) else r.get('title', ''), 
            r['date']
        ), 
        axis=1
    )
    
    # 7. Generowanie kluczy dla REPORT
    report_df['__merge_key'] = report_df.apply(
        lambda r: create_key(
            r.get(media_col, ''), 
            r.get(title_col, ''), 
            r.get(date_col, '')
        ), 
        axis=1
    )
    report_df['__merge_key_main'] = report_df.apply(
        lambda r: create_key_main_domain(
            r.get(media_col, ''), 
            r.get(title_col, ''), 
            r.get(date_col, '')
        ), 
        axis=1
    )
    
    # 8. Mapowanie wartości - PR Value i inne kolumny z raportu
    pr_map_full = dict(zip(report_df['__merge_key'], report_df[pr_col]))
    pr_map_main = dict(zip(report_df['__merge_key_main'], report_df[pr_col]))
    
    # Mapy dla dodatkowych kolumn z raportu
    extra_cols_to_map = ['Uploaded', 'Published', 'Media Type', 'Media Grade', 'Country']
    extra_maps_full = {}
    extra_maps_main = {}
    for col in extra_cols_to_map:
        if col in report_df.columns:
            extra_maps_full[col] = dict(zip(report_df['__merge_key'], report_df[col]))
            extra_maps_main[col] = dict(zip(report_df['__merge_key_main'], report_df[col]))
    
    # Najpierw próbuj pełny klucz
    clean_df['PR Value'] = clean_df['__merge_key'].map(pr_map_full)
    for col, col_map in extra_maps_full.items():
        clean_df[col] = clean_df['__merge_key'].map(col_map)
    
    # Dla tych co nie mają matcha - użyj fallback z główną domeną
    mask_no_match = clean_df['PR Value'].isna()
    clean_df.loc[mask_no_match, 'PR Value'] = clean_df.loc[mask_no_match, '__merge_key_main'].map(pr_map_main)
    for col, col_map in extra_maps_main.items():
        clean_df.loc[mask_no_match, col] = clean_df.loc[mask_no_match, '__merge_key_main'].map(col_map)
    
    # 9. Dla wciąż brakujących - wstaw [ERROR] (nie 0!)
    mask_still_no_match = clean_df['PR Value'].isna()
    clean_df['PR Value'] = clean_df['PR Value'].astype(object)  # Żeby móc wstawić string
    clean_df.loc[mask_still_no_match, 'PR Value'] = '[ERROR]'
    
    # 10. Usuń kolumny pomocnicze
    clean_df.drop(columns=['__merge_key', '__merge_key_main'], inplace=True)
    report_df.drop(columns=['__merge_key', '__merge_key_main'], inplace=True)
    
    return clean_df

def clean_json_response(raw_resp):
    try:
        start = raw_resp.find('{')
        end = raw_resp.rfind('}') + 1
        if start != -1 and end != -1:
            clean_str = raw_resp[start:end]
            return json.loads(clean_str)
        else: return None
    except: return None

def call_openai_single(prompt, key, model):
    url = "https://api.openai.com/v1/chat/completions"
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {key}"}
    payload = {
        "model": model,
        "messages": [{"role": "system", "content": "You are a Data Analyst."}, {"role": "user", "content": prompt}],
        "temperature": 0.1
    }
    
    for attempt in range(3):
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=25)
            if resp.status_code == 200:
                return resp.json()['choices'][0]['message']['content']
            elif resp.status_code == 429:
                time.sleep(2 * (attempt + 1))
                continue
            else:
                return f"[API_ERROR: {resp.status_code}]"
        except Exception as e:
            if attempt == 2: return f"[CONN_ERR: {str(e)[:20]}]"
            time.sleep(1)
    return "[TIMEOUT]"

def call_openai_vision(prompt, img_url, key):
    url = "https://api.openai.com/v1/chat/completions"
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {key}"}
    payload = {
        "model": "gpt-4o",
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": img_url, "detail": "low"}}
                ]
            }
        ],
        "max_tokens": 50
    }
    
    for attempt in range(3):
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=25)
            if resp.status_code == 200:
                return resp.json()['choices'][0]['message']['content']
            elif resp.status_code == 429:
                time.sleep(2 * (attempt + 1))
                continue
            else:
                return f"[API_ERROR: {resp.status_code}]"
        except Exception as e:
            if attempt == 2: return f"[CONN_ERR: {str(e)[:20]}]"
            time.sleep(1)
    return "[TIMEOUT]"

def analyze_row_with_ai(row, api_key):
    needs_div = not has_value(row['Division'])
    needs_prod = not has_value(row['Product'])
    needs_excl = not has_value(row['Exclusive'])
    needs_quote = not has_value(row['Quote'])
    needs_photo = not has_value(row['Photo'])
    needs_eng = not has_value(row['ENG Title'])
    needs_esg = not has_value(row['ESG'])
    needs_mz = not has_value(row['M/Z'])
    
    if not any([needs_div, needs_prod, needs_excl, needs_quote, needs_photo, needs_eng, needs_esg, needs_mz]):
        return None

    url = row.get('Links', '')
    scraped = scrape_article_data(url) if url else {"text": "", "image_url": None}
    text_content = scraped.get('text', '')
    img_url = scraped.get('image_url')
    source_text = text_content
    source_note = ""
    orig_title = str(row.get('title', ''))  # Zmiana z 'tytul' na 'title'

    if not source_text or len(source_text) < 50:
        source_text = orig_title
        source_note = "[TITLE ONLY] "
        
    updates = {}

    if any([needs_div, needs_prod, needs_excl, needs_quote, needs_eng, needs_esg, needs_mz]):
        current_div = row.get('Division', '') if has_value(row['Division']) else ""
        current_prod = row.get('Product', '') if has_value(row['Product']) else ""
        
        constraint_txt = ""
        if current_div: constraint_txt += f" CONSTRAINT: Division is FIXED to '{current_div}'. Select Product ONLY from its list."
        if current_prod: constraint_txt += f" CONSTRAINT: Product is FIXED to '{current_prod}'. Infer Division from it."

        if not source_text or len(source_text) < 5:
             err_msg = "[NO_CONTENT]"
             if needs_div: updates['Division'] = err_msg
             if needs_prod: updates['Product'] = err_msg
             if needs_excl: updates['Exclusive'] = err_msg
             if needs_quote: updates['Quote'] = err_msg
             if needs_eng: updates['ENG Title'] = err_msg
             if needs_esg: updates['ESG'] = "No"
             if needs_mz: updates['M/Z'] = "No"
        else:
            prompt = f"""
            Analyze article about LG Electronics. {source_note}
            Original Title: "{orig_title}"
            Product Map: {json.dumps(PRODUCT_RULES)}
            
            Rules:
            1. Identify Division and Product. {constraint_txt}
            2. If NOT about LG (e.g. Chem, Solar), Division='Corporate', Product='Others'.
            3. Exclusive: ONLY two values allowed: 'Exclusive' or '66'.
               - Default to 'Exclusive' in most cases.
               - Use '66' ONLY when LG is barely mentioned (just a brief reference, not the main topic).
               - When in doubt, choose 'Exclusive'.
            4. Quote: Extract 1 relevant sentence (max 150 chars) AND TRANSLATE it to US English.
               CONSTRAINT: If the quote contains "LG", keep "LG".
               CONSTRAINT: Do NOT include any punctuation at the end (no dots, commas, exclamation marks etc.)
            5. Translate 'Original Title' to US English (field: 'EngTitle').
               CONSTRAINT: If 'Original Title' contains "LG", the 'EngTitle' MUST also contain "LG".
            6. ESG: Return 'Yes' if article mentions ANY of these: heat pumps, air conditioners, chillers, 
               water heaters, HVAC products, refrigerators, washing machines, washtower, clothes dryers, 
               energy efficiency, ecology, water saving, product freshness. Otherwise 'No'.
            7. MZ (Entertainment/Gaming): Return 'Yes' if article mentions ANY of these: monitors, gaming monitors, 
               games, TV/television, soundbar, projector, headphones, speakers, xboom. Otherwise 'No'.
            
            Return JSON: {{ "Division": "...", "Product": "...", "Exclusive": "...", "Quote": "...", "EngTitle": "...", "ESG": "Yes/No", "MZ": "Yes/No" }}
            Text: {source_text[:2500]}
            """
            
            raw_resp = call_openai_single(prompt, api_key, "gpt-4o-mini")
            data = clean_json_response(raw_resp)
            
            if data:
                if needs_div: updates['Division'] = enforce_strict_rules("Division", data.get('Division', ''))
                if needs_prod: updates['Product'] = enforce_strict_rules("Product", data.get('Product', ''), updates.get('Division', current_div))
                if needs_excl: updates['Exclusive'] = enforce_strict_rules("Exclusive", data.get('Exclusive', ''))
                if needs_quote: updates['Quote'] = data.get('Quote', '')
                if needs_eng: updates['ENG Title'] = data.get('EngTitle', '')
                if needs_esg: updates['ESG'] = "Yes" if str(data.get('ESG', '')).lower() == 'yes' else "No"
                if needs_mz: updates['M/Z'] = "Yes" if str(data.get('MZ', '')).lower() == 'yes' else "No"
            else:
                err_frag = f"[JSON_ERR: {raw_resp[:20]}]"
                if needs_div: updates['Division'] = err_frag
                if needs_prod: updates['Product'] = err_frag
                if needs_excl: updates['Exclusive'] = err_frag
                if needs_quote: updates['Quote'] = err_frag
                if needs_eng: updates['ENG Title'] = err_frag
                if needs_esg: updates['ESG'] = "No"
                if needs_mz: updates['M/Z'] = "No"

    if needs_photo:
        if img_url:
            vision_prompt = "What is in this image related to LG? Return ONLY one string: 'LGE logo', 'product', 'personnel', or 'None'."
            raw_vis = call_openai_vision(vision_prompt, img_url, api_key)
            clean_vis = raw_vis.replace("'", "").replace('"', '').replace(".", "").strip()
            updates['Photo'] = enforce_strict_rules("Photo", clean_vis)
        else:
            updates['Photo'] = "None"

    if not updates: return None
    return {"index": row.name, "changes": updates}

# --- AGGRID HELPER ---
def prepare_aggrid_data(df):
    # Clean title - używamy ENG Title jeśli jest, w przeciwnym razie title
    df['_clean_title'] = df.apply(
        lambda r: clean_text(r['ENG Title'] if has_value(r['ENG Title']) else r['title'], TITLE_MAX_LEN), 
        axis=1
    )
    # Clean quote
    df['_clean_quote'] = df['Quote'].apply(lambda x: clean_text(x, QUOTE_MAX_LEN))
    # ID Match do mergowania
    df['_ID_MATCH'] = df.apply(generate_id_match, axis=1)
    return df

def prepare_final_export(df):
    """Przygotowuje DataFrame do eksportu w finalnym formacie."""
    export_df = df.copy()
    
    # Upewnij się, że wszystkie kolumny finalne istnieją
    for col in FINAL_OUTPUT_ORDER:
        if col not in export_df.columns:
            export_df[col] = ""
    
    # Usuń kolumny wewnętrzne (zaczynające się od _)
    internal_cols = [c for c in export_df.columns if c.startswith('_')]
    export_df = export_df.drop(columns=internal_cols, errors='ignore')
    
    # Zwróć tylko kolumny w odpowiedniej kolejności
    final_cols = [c for c in FINAL_OUTPUT_ORDER if c in export_df.columns]
    return export_df[final_cols]


def prepare_raw_data_export(df):
    """
    Przygotowuje plik LGePR_raw_data.
    Kolumny: Uploaded, Published, Media, Media Type, Media Grade, Headline, Country, Division, Product, PR Value, ESG, M/Z
    """
    export_df = df.copy()
    
    # Mapowanie kolumn
    raw_data = pd.DataFrame()
    raw_data['Uploaded'] = export_df.get('Uploaded', '')
    raw_data['Published'] = export_df.get('Published', '')
    raw_data['Media'] = export_df.get('source', '')
    raw_data['Media Type'] = export_df.get('Media Type', 'ONLINE')
    raw_data['Media Grade'] = export_df.get('Media Grade', '')
    raw_data['Headline'] = export_df.get('ENG Title', '')
    raw_data['Country'] = export_df.get('Country', 'Poland')
    raw_data['Division'] = export_df.get('Division', '')
    raw_data['Product'] = export_df.get('Product', '')
    raw_data['PR Value'] = export_df.get('PR Value', '')
    raw_data['ESG'] = export_df.get('ESG', '')
    raw_data['M/Z'] = export_df.get('M/Z', '')
    
    return raw_data


def prepare_reach_export(df):
    """
    Przygotowuje plik Reach.
    Kolumny: Division, Product, Medium, Source, Title.pl, Title.eng, Reach, Date, ESG, MZ
    """
    export_df = df.copy()
    
    reach_data = pd.DataFrame()
    reach_data['Division'] = export_df.get('Division', '')
    reach_data['Product'] = export_df.get('Product', '')
    # Medium - ONLINE lub PRINT (z Media Type)
    reach_data['Medium'] = export_df.get('Media Type', 'ONLINE')
    reach_data['Source'] = export_df.get('source', '')
    reach_data['Title.pl'] = export_df.get('title', '')  # Oryginalny polski tytuł
    reach_data['Title.eng'] = export_df.get('ENG Title', '')  # Angielski tytuł
    reach_data['Reach'] = export_df.get('reach', '')
    reach_data['Date'] = export_df.get('Published', '')  # Data publikacji
    reach_data['ESG'] = export_df.get('ESG', '')
    reach_data['MZ'] = export_df.get('M/Z', '')
    
    return reach_data

def calculate_statistics(df):
    """
    Oblicza statystyki PR Value i Reach po merge.
    Zwraca dwie tabele: główną (PR Value + Reach) i szczegółową (tylko PR Value).
    """
    
    # Konwertuj kolumny na numeric (mogą być stringi ze spacjami lub [ERROR])
    def safe_to_numeric(val):
        if pd.isna(val) or val == '[ERROR]':
            return 0
        val_str = str(val).replace(' ', '').replace('\u00a0', '')
        try:
            return float(val_str)
        except:
            return 0
    
    df['_pr_numeric'] = df['PR Value'].apply(safe_to_numeric)
    df['_reach_numeric'] = df['reach'].apply(safe_to_numeric)
    
    # === TABELA 1: Główna (PR Value + Reach) ===
    table1_data = {
        'Kategoria': [
            'Corporate',
            'HS',
            'ES',
            'MS (LCD TV + OLED TV + Others)',
            'MS Audio',
            'MS (Signage + PC + Projector + Monitor)'
        ],
        'PR Value': [
            df[df['Division'] == 'Corporate']['_pr_numeric'].sum(),
            df[df['Division'] == 'HS']['_pr_numeric'].sum(),
            df[df['Division'] == 'ES']['_pr_numeric'].sum(),
            df[(df['Division'] == 'MS') & (df['Product'].isin(['LCD TV', 'OLED TV', 'Others']))]['_pr_numeric'].sum(),
            df[(df['Division'] == 'MS') & (df['Product'] == 'Audio')]['_pr_numeric'].sum(),
            df[(df['Division'] == 'MS') & (df['Product'].isin(['Signage', 'PC', 'Projector', 'Monitor']))]['_pr_numeric'].sum(),
        ],
        'Reach': [
            df[df['Division'] == 'Corporate']['_reach_numeric'].sum(),
            df[df['Division'] == 'HS']['_reach_numeric'].sum(),
            df[df['Division'] == 'ES']['_reach_numeric'].sum(),
            df[(df['Division'] == 'MS') & (df['Product'].isin(['LCD TV', 'OLED TV', 'Others']))]['_reach_numeric'].sum(),
            df[(df['Division'] == 'MS') & (df['Product'] == 'Audio')]['_reach_numeric'].sum(),
            df[(df['Division'] == 'MS') & (df['Product'].isin(['Signage', 'PC', 'Projector', 'Monitor']))]['_reach_numeric'].sum(),
        ]
    }
    table1 = pd.DataFrame(table1_data)
    
    # === TABELA 2: Szczegółowa (tylko PR Value) - POZIOMA ===
    table2_data = {
        'HS': [df[df['Division'] == 'HS']['_pr_numeric'].sum()],
        'ES': [df[df['Division'] == 'ES']['_pr_numeric'].sum()],
        'MS (LCD TV + OLED TV + Others)': [df[(df['Division'] == 'MS') & (df['Product'].isin(['LCD TV', 'OLED TV', 'Others']))]['_pr_numeric'].sum()],
        'MS Audio': [df[(df['Division'] == 'MS') & (df['Product'] == 'Audio')]['_pr_numeric'].sum()],
        'MS (Signage + PC + Projector + Monitor)': [df[(df['Division'] == 'MS') & (df['Product'].isin(['Signage', 'PC', 'Projector', 'Monitor']))]['_pr_numeric'].sum()],
        'Corporate': [df[df['Division'] == 'Corporate']['_pr_numeric'].sum()],
        'ESG (Yes)': [df[df['ESG'] == 'Yes']['_pr_numeric'].sum()],
    }
    table2 = pd.DataFrame(table2_data)
    
    # Usuń kolumny pomocnicze
    df.drop(columns=['_pr_numeric', '_reach_numeric'], inplace=True, errors='ignore')
    
    return table1, table2


# ─────────────────────────────────────────────
# 6. GŁÓWNA APLIKACJA
# ─────────────────────────────────────────────
def main():
    st.title("🧹 LGePR Data Cleaner v12.7")

    if not AGGRID_AVAILABLE:
        st.error("❌ Brak biblioteki streamlit-aggrid. Zainstaluj ją komendą: pip install streamlit-aggrid")
        st.stop()

    if 'config_loaded' not in st.session_state:
        secret_key, secret_media = get_cloud_config()
        st.session_state.saved_api_key = secret_key
        st.session_state.media_list = secret_media
        st.session_state.config_loaded = True
        st.session_state.step = 0  # 0 = ekran startowy
        st.session_state.process = None  # 'clean' lub 'merge'
        st.session_state.df_work = None
        st.session_state.ai_proposals = None
        st.session_state.grid_key_suffix = 0 

    with st.sidebar:
        st.header("Ustawienia")
        if st.session_state.saved_api_key:
            st.success("✅ Klucz API (Secrets)")
            active_key = st.session_state.saved_api_key
        else:
            active_key = st.text_input("Klucz API (Tymczasowy)", type="password")
        
        st.divider()
        st.header("Media")
        if st.session_state.media_list:
            st.success(f"✅ Baza mediów (Secrets): {len(st.session_state.media_list)}")
            
            # Formularz dodawania nowego medium
            with st.expander("➕ Dodaj medium do bazy"):
                new_media = st.text_input("Nazwa domeny (np. nowastrona.pl)", key="new_media_input")
                if st.button("Dodaj do bazy", key="add_media_btn"):
                    if new_media and new_media.strip():
                        clean_media = normalize_domain(new_media).lower()
                        if clean_media:
                            st.session_state.media_list.add(clean_media)
                            st.success(f"✅ Dodano: {clean_media}")
                            st.info(f"Baza mediów: {len(st.session_state.media_list)} pozycji")
                        else:
                            st.error("Nieprawidłowa nazwa domeny")
                    else:
                        st.warning("Wpisz nazwę domeny")
        else:
            st.warning("Brak listy mediów w Secrets. Użyj pliku tymczasowego.")

    curr = st.session_state.step
    process = st.session_state.process

    # ===== EKRAN STARTOWY =====
    if curr == 0:
        st.markdown("## 🏠 Wybierz proces")
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📤 Upload, Analiza, Weryfikacja")
            st.markdown("Wgraj plik roboczy, przeanalizuj AI i zweryfikuj dane.")
            if st.button("▶️ Rozpocznij proces Clean", type="primary", use_container_width=True):
                st.session_state.process = 'clean'
                st.session_state.step = 1
                st.rerun()
        
        with col2:
            st.markdown("### 🔗 Merge z PR Value")
            st.markdown("Połącz gotowy plik Clean z raportem PR Value.")
            if st.button("▶️ Przejdź do Merge", type="primary", use_container_width=True):
                st.session_state.process = 'merge'
                st.session_state.step = 4
                st.rerun()
        
        return  # Koniec ekranu startowego

    # ===== PASEK KROKÓW DLA PROCESU CLEAN =====
    if process == 'clean':
        s1, s2, s3 = st.columns(3)
        
        with s1:
            if curr == 1:
                st.info("1. Upload")
            else:
                st.markdown("1. Upload")
        with s2:
            if curr == 2:
                st.info("2. Analiza AI")
            else:
                st.markdown("2. Analiza AI")
        with s3:
            if curr == 3:
                st.info("3. Weryfikacja")
            else:
                st.markdown("3. Weryfikacja")
        
        st.divider()
        
        # Przycisk powrotu do ekranu startowego
        if st.button("🏠 Powrót do menu głównego"):
            st.session_state.step = 0
            st.session_state.process = None
            st.session_state.df_work = None
            st.rerun()

    # ===== PASEK DLA PROCESU MERGE =====
    elif process == 'merge':
        st.info("🔗 Merge z PR Value")
        st.divider()
        
        if st.button("🏠 Powrót do menu głównego"):
            st.session_state.step = 0
            st.session_state.process = None
            st.rerun()

    # ===== KROK 1: UPLOAD =====
    if curr == 1:
        f = st.file_uploader("Wgraj plik roboczy (.xlsx)", type=['xlsx', 'xlsm'], key="upload_step1")
        if f:
            try:
                wb = openpyxl.load_workbook(f, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                sh = st.selectbox("Arkusz:", sheets)
                if st.button("🚀 Załaduj i Pokaż", type="primary"):
                    f.seek(0)
                    df = extract_specific_columns(f, sh, st.session_state.media_list)
                    st.session_state.df_work = df
                    st.session_state.grid_key_suffix += 1 
                    st.success(f"Wczytano {len(df)} wierszy.")
                    st.rerun()
            except Exception as e:
                st.error(f"Błąd pliku: {e}")
        
        if st.session_state.df_work is not None:
            st.markdown(f"### 📄 Podgląd danych (Cały plik: {len(st.session_state.df_work)} wierszy)")
            st.dataframe(st.session_state.df_work, use_container_width=True, height=500)
            col_btn, _ = st.columns([1, 4])
            with col_btn:
                if st.button("Przejdź do Analizy →", type="primary"):
                    st.session_state.step = 2
                    st.rerun()

    elif curr == 2:
        df = st.session_state.df_work
        st.markdown("### 🧠 Analiza treści, obrazu i tłumaczenie")
        st.info("AI przeanalizuje linki, uzupełni pola, pobierze zdjęcia i PRZETŁUMACZY (Title/Quote) na US English (zachowując 'LG').")
        
        c1, c2 = st.columns([1, 3])
        with c1:
            run_analysis = st.button("▶️ Uruchom Pełną Analizę", type="primary", disabled=not active_key)
        
        if run_analysis:
            progress_bar = st.progress(0)
            status_text = st.empty()
            proposals = []
            total = len(df)
            
            for i, row in df.iterrows():
                status_text.text(f"Analizuję wiersz {i+1}/{total}: {str(row['title'])[:30]}...")
                update = analyze_row_with_ai(row, active_key)
                if update:
                    proposals.append(update)
                progress_bar.progress((i + 1) / total)
            
            status_text.success("Analiza zakończona!")
            if proposals:
                st.session_state.ai_proposals = proposals
                st.rerun()
            else:
                st.warning("Wszystko wygląda na uzupełnione lub brak danych do analizy.")

        if st.session_state.ai_proposals:
            st.divider()
            st.markdown(f"**Znaleziono {len(st.session_state.ai_proposals)} sugestii zmian.**")
            
            prop_data = []
            for p in st.session_state.ai_proposals:
                for k, v in p['changes'].items():
                    prop_data.append({
                        "Idx": p['index'],
                        "Wiersz": p['index'] + 2,
                        "Kolumna": k,
                        "Obecnie": df.at[p['index'], k],
                        "Sugestia AI": v
                    })
            
            edited_props = st.data_editor(
                pd.DataFrame(prop_data),
                use_container_width=True,
                disabled=["Wiersz", "Kolumna", "Obecnie"],
                column_config={"Idx": None},
                hide_index=True
            )
            
            if st.button("✅ Zatwierdź zmiany i przejdź dalej"):
                for _, row_p in edited_props.iterrows():
                    st.session_state.df_work.at[row_p['Idx'], row_p['Kolumna']] = row_p['Sugestia AI']
                st.session_state.ai_proposals = None
                st.session_state.step = 3
                st.rerun()
        else:
            col_back, col_next = st.columns([1, 1])
            with col_back:
                if st.button("← Wróć do Upload"):
                    st.session_state.step = 1
                    st.rerun()
            with col_next:
                if st.button("Pomiń / Dalej →"):
                    st.session_state.step = 3
                    st.rerun()

    elif curr == 3:
        st.markdown("### 🔍 Weryfikacja i Edycja (AgGrid Live)")
        
        df_prepared = prepare_aggrid_data(st.session_state.df_work)
        cols = [c for c in FINAL_OUTPUT_ORDER if c in df_prepared.columns]
        # Dodaj kolumny wewnętrzne potrzebne do weryfikacji
        if '_media_status' in df_prepared.columns:
            cols.append('_media_status')
        if '_orig_date' in df_prepared.columns:
             cols.append('_orig_date')
        
        custom_css = {
            ".cell-error": {
                "background-color": "#ffcccc !important",
                "color": "darkred !important",
                "font-weight": "bold !important"
            }
        }
        
        js_division = JsCode(f"""
        {{
            'cell-error': function(params) {{
                const allowed = {json.dumps(VALIDATION_RULES["Division"])};
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                return !allowed.includes(val);
            }}
        }}
        """)

        js_product = JsCode(f"""
        {{
            'cell-error': function(params) {{
                const map = {json.dumps(PRODUCT_RULES)};
                let div = params.data.Division;
                if (!div) div = "";
                div = div.toString().trim();
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                let allowed = [];
                if (map[div]) {{ allowed = map[div]; }} else {{ Object.values(map).forEach(arr => allowed.push(...arr)); }}
                return !allowed.includes(val);
            }}
        }}
        """)
        
        js_photo = JsCode(f"""
        {{
            'cell-error': function(params) {{
                const allowed = {json.dumps(VALIDATION_RULES["Photo"])};
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                return !allowed.includes(val);
            }}
        }}
        """)
        
        js_exclusive = JsCode(f"""
        {{
            'cell-error': function(params) {{
                const allowed = {json.dumps(VALIDATION_RULES["Exclusive"])};
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                return !allowed.includes(val);
            }}
        }}
        """)
        
        js_lg = JsCode(f"""
        {{
            'cell-error': function(params) {{
                const allowed = {json.dumps(VALIDATION_RULES["LG"])};
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                return !allowed.includes(val);
            }}
        }}
        """)
        
        js_esg = JsCode(f"""
        {{
            'cell-error': function(params) {{
                const allowed = {json.dumps(VALIDATION_RULES["ESG"])};
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                return !allowed.includes(val);
            }}
        }}
        """)
        
        js_mz = JsCode(f"""
        {{
            'cell-error': function(params) {{
                const allowed = {json.dumps(VALIDATION_RULES["M/Z"])};
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                return !allowed.includes(val);
            }}
        }}
        """)
        
        js_media = JsCode("""
        {
            'cell-error': function(params) {
                let val = params.value;
                if (val === null || val === undefined) val = "";
                val = val.toString().trim();
                return val === 'BRAK';
            }
        }
        """)

        gb = GridOptionsBuilder.from_dataframe(df_prepared[cols])
        gb.configure_default_column(editable=True, resizable=True, wrapText=True, autoHeight=True)
        gb.configure_column('_ID_MATCH', editable=False, hide=True)
        
        if '_orig_date' in df_prepared.columns:
            gb.configure_column('_orig_date', hide=True)
        
        gb.configure_column('Division', cellClassRules=js_division)
        gb.configure_column('Product', cellClassRules=js_product)
        gb.configure_column('Photo', cellClassRules=js_photo)
        gb.configure_column('Exclusive', cellClassRules=js_exclusive)
        gb.configure_column('LG', cellClassRules=js_lg)
        gb.configure_column('ESG', cellClassRules=js_esg)
        gb.configure_column('M/Z', cellClassRules=js_mz)
        gb.configure_column('_media_status', cellClassRules=js_media)

        gb.configure_grid_options(domLayout='normal', height=600)
        gridOptions = gb.build()

        current_grid_key = f"editor_grid_{st.session_state.grid_key_suffix}"
        
        grid_response = AgGrid(
            df_prepared[cols], 
            gridOptions=gridOptions, 
            custom_css=custom_css, 
            allow_unsafe_jscode=True, 
            update_mode=GridUpdateMode.VALUE_CHANGED,
            data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
            fit_columns_on_grid_load=False,
            enable_enterprise_modules=False,
            key=current_grid_key, 
            reload_data=False 
        )

        updated_df = pd.DataFrame(grid_response['data'])
        
        if not updated_df.equals(st.session_state.df_work):
             st.session_state.df_work = updated_df

        err_count = 0
        for i, row in updated_df.iterrows():
            div = str(row.get('Division', '')).strip()
            if div not in VALIDATION_RULES['Division']:
                err_count += 1
            else:
                allowed = PRODUCT_RULES.get(div, [])
                if str(row.get('Product', '')).strip() not in allowed:
                    err_count += 1
            if str(row.get('Photo', '')).strip() not in VALIDATION_RULES['Photo']:
                err_count += 1
            if str(row.get('Exclusive', '')).strip() not in VALIDATION_RULES['Exclusive']:
                err_count += 1
            if str(row.get('LG', '')).strip() not in VALIDATION_RULES['LG']:
                err_count += 1
            if str(row.get('ESG', '')).strip() not in VALIDATION_RULES['ESG']:
                err_count += 1
            if str(row.get('M/Z', '')).strip() not in VALIDATION_RULES['M/Z']:
                err_count += 1
            if str(row.get('_media_status', '')).strip() == 'BRAK':
                err_count += 1

        if err_count > 0:
            st.warning(f"⚠️ Znaleziono ok. {err_count} pól do poprawy (podświetlone na czerwono).")
        else:
            st.success("✅ Wszystkie pola wyglądają poprawnie!")

        col_d1, col_d2 = st.columns(2)
        with col_d1:
            if st.button("← Wróć do Analizy"):
                st.session_state.step = 2
                st.rerun()
        
        with col_d2:
            # Przygotuj dane do eksportu w finalnym formacie
            export_df = prepare_final_export(st.session_state.df_work)
            b = io.BytesIO()
            with pd.ExcelWriter(b, engine='xlsxwriter') as w:
                export_df.to_excel(w, sheet_name='Report', index=False)
            
            st.download_button(
                label="⬇️ Pobierz Czysty Plik", 
                data=b.getvalue(), 
                file_name="LGePR_Clean.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                type="primary"
            )

    elif curr == 4:
        st.markdown("### 🔗 Łączenie z Raportem PR Value")
        c1, c2 = st.columns(2)
        with c1:
            f_clean = st.file_uploader("1. Twój Plik Czysty", type=['xlsx'], key="upload_clean")
        with c2:
            f_report = st.file_uploader("2. Raport z systemu (z PR Value)", type=['xlsx'], key="upload_report")
        
        if f_clean and f_report:
            if st.button("🔗 Połącz Pliki", type="primary"):
                try:
                    df_c = pd.read_excel(f_clean)
                    df_r = pd.read_excel(f_report)
                    df_final = merge_datasets(df_c, df_r)
                    st.success("Połączono pomyślnie!")
                    
                    # === STATYSTYKI ===
                    st.divider()
                    st.markdown("### 📊 Statystyki")
                    
                    # Łączna liczba publikacji
                    total_publications = len(df_final)
                    st.metric("📰 Łączna liczba publikacji", total_publications)
                    
                    table1, table2 = calculate_statistics(df_final)
                    
                    col_t1, col_t2 = st.columns(2)
                    
                    with col_t1:
                        st.markdown("**Tabela 1: PR Value + Reach**")
                        st.dataframe(table1, use_container_width=True, hide_index=True)
                    
                    with col_t2:
                        st.markdown("**Tabela 2: PR Value (szczegółowy)**")
                        st.dataframe(table2, use_container_width=True, hide_index=True)
                    
                    st.divider()
                    
                    # Podgląd danych
                    st.markdown("### 📄 Podgląd połączonych danych")
                    st.dataframe(df_final[['source', 'title', 'Division', 'Product', 'PR Value', 'reach']].head(10), use_container_width=True)
                    
                    st.divider()
                    st.markdown("### 📥 Pobierz pliki")
                    
                    col_dl1, col_dl2, col_dl3 = st.columns(3)
                    
                    # 1. FINAL RAPORT
                    with col_dl1:
                        export_final = prepare_final_export(df_final)
                        b_fin = io.BytesIO()
                        with pd.ExcelWriter(b_fin, engine='xlsxwriter') as w:
                            export_final.to_excel(w, sheet_name='Report', index=False)
                        st.download_button(
                            "⬇️ LGePR_FINAL",
                            b_fin.getvalue(),
                            f"LGePR_FINAL_{datetime.now().strftime('%d%m')}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True
                        )
                    
                    # 2. RAW DATA
                    with col_dl2:
                        export_raw = prepare_raw_data_export(df_final)
                        b_raw = io.BytesIO()
                        with pd.ExcelWriter(b_raw, engine='xlsxwriter') as w:
                            export_raw.to_excel(w, sheet_name='Raw Data', index=False)
                        st.download_button(
                            "⬇️ LGePR_raw_data",
                            b_raw.getvalue(),
                            f"LGePR_raw_data_{datetime.now().strftime('%d%m')}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="secondary",
                            use_container_width=True
                        )
                    
                    # 3. REACH
                    with col_dl3:
                        export_reach = prepare_reach_export(df_final)
                        b_reach = io.BytesIO()
                        with pd.ExcelWriter(b_reach, engine='xlsxwriter') as w:
                            export_reach.to_excel(w, sheet_name='Reach', index=False)
                        st.download_button(
                            "⬇️ Reach",
                            b_reach.getvalue(),
                            f"Reach_{datetime.now().strftime('%d%m')}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="secondary",
                            use_container_width=True
                        )
                        
                except Exception as e:
                    st.error(f"Błąd łączenia: {e}")
        
        if st.button("← Wróć do Weryfikacji"):
            st.session_state.step = 3
            st.rerun()

if __name__ == "__main__":
    main()
